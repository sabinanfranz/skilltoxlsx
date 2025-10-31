# -*- coding: utf-8 -*-
import streamlit as st
import streamlit.components.v1 as components
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Tuple
import json
import re
import base64

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ==========================
# 상수 / 경로
# ==========================
APP_DIR = Path(__file__).parent
TEMPLATE_DIR = APP_DIR / "templates"
DEFAULT_TEMPLATE_NONTRACK = "Non Track_Paper Interview_상위조직명_직무명(포맷).xlsx"
DEFAULT_TEMPLATE_TRACK    = "Track_Paper Interview_상위조직명_직무명(포맷).xlsx"

# Non Track 쓰기 범위
TASK_START_ROW_NT, TASK_END_ROW_NT   = 5, 14   # Task: A(이름), C(설명)
SKILL_START_ROW_NT, SKILL_END_ROW_NT = 5, 11   # Skill: A/B/D/F

# Track 쓰기 범위 (규칙 동일)
TASK_ROW_START_T, TASK_ROW_END_T   = 5, 14
SKILL_ROW_START_T, SKILL_ROW_END_T = 5, 11
TASK_TEMPLATE_SHEET_T  = "Task"
SKILL_TEMPLATE_SHEET_T = "Skill"
TRACK_TITLE_RANGE_T    = "D1:D2"  # 트랙명 표기 영역

# ==========================
# 공통: 텍스트 정리(마커 제거)
# ==========================
# [cite: ...]
CITE_PATTERN = re.compile(r'\s*\[\s*cite\s*:\s*.*?\]\s*', flags=re.IGNORECASE | re.DOTALL)
# (Source ...)
SOURCE_PAREN_PATTERN = re.compile(r'\s*\(\s*source[^)]*\)\s*', flags=re.IGNORECASE)

def strip_markers(text: Any) -> str:
    """[cite: ...], (Source ...) 제거 + 공백 정리"""
    if text is None:
        return ""
    s = str(text)
    s = CITE_PATTERN.sub(" ", s)
    s = SOURCE_PAREN_PATTERN.sub(" ", s)
    s = re.sub(r"[ \t]+", " ", s).strip()
    return s

# ==========================
# 공통: 파일명 유틸
# ==========================
INVALID_WIN_CHARS = r'<>:"/\\|?*'
INVALID_WIN_PATTERN = re.compile(f"[{re.escape(INVALID_WIN_CHARS)}]+")

def sanitize_filename_component(s: str, fallback: str = "untitled") -> str:
    if not s:
        return fallback
    s = INVALID_WIN_PATTERN.sub(" ", s).strip().strip(".")
    return s if s else fallback

# ==========================
# Non Track 파서/로직
# ==========================
def title_tokens_nt(stem: str) -> List[str]:
    return [t.strip() for t in stem.split("_") if t.strip()]

def is_trailing_excluded_nt(token: str) -> bool:
    t = token.lower().replace(" ", "")
    return t in {"skill", "hc제외"}

def parse_org_role_from_filename_nt(filename: str) -> Tuple[str, str, str]:
    """{상위조직명} = 첫 토큰, {직무명} = 두 번째~끝(뒤에서 skill/HC 제외 제거), 표시/파일명 둘 다 '공백' 연결"""
    stem = Path(filename).stem
    toks = title_tokens_nt(stem)
    if not toks:
        return "unknown", "", ""
    org = toks[0]
    end = len(toks)
    while end > 1 and is_trailing_excluded_nt(toks[end - 1]):
        end -= 1
    role_tokens = toks[1:end] if end > 1 else toks[1:]
    role_display = " ".join(role_tokens)
    role_for_filename = " ".join(role_tokens)
    return org, role_display, role_for_filename

def with_wrap(cell):
    a = cell.alignment or Alignment()
    return Alignment(
        horizontal=a.horizontal,
        vertical=a.vertical,
        text_rotation=a.text_rotation,
        wrap_text=True,
        shrink_to_fit=a.shrink_to_fit,
        indent=a.indent
    )

def set_text(ws, coord: str, text: str, wrap: bool = True):
    cell = ws[coord]
    cell.value = text
    if wrap:
        cell.alignment = with_wrap(cell)

def load_json_from_txt_bytes(b: bytes) -> Dict[str, Any]:
    """TXT에 전후 텍스트가 섞여 있어도 {} 블록만 추출 시도"""
    txt = b.decode("utf-8-sig", errors="ignore")
    try:
        return json.loads(txt)
    except json.JSONDecodeError:
        start = txt.find("{")
        end = txt.rfind("}")
        if start != -1 and end != -1 and start < end:
            return json.loads(txt[start:end+1])
        raise

def collect_tasks_nt(obj: Dict[str, Any]) -> List[Dict[str, Any]]:
    return obj.get("tasks") or []

def iter_skills_nt(obj: Dict[str, Any]):
    skills = obj.get("skills") or []
    for item in skills:
        if isinstance(item, dict) and "skill" in item:
            s = item.get("skill") or {}
            name = s.get("name", "")
            definition = s.get("definition", "")
            stack = s.get("tech_stack", {})
            related = item.get("related_tasks") or s.get("related_tasks") or []
        else:
            s = item if isinstance(item, dict) else {}
            name = s.get("name", "")
            definition = s.get("definition", "")
            stack = s.get("tech_stack", {})
            related = s.get("related_tasks") or []
        yield {"name": name, "definition": definition, "tech_stack": stack, "related_tasks": related}

def normalize_list(val) -> List[str]:
    if val is None:
        return []
    if isinstance(val, (list, tuple, set)):
        return [str(x).strip() for x in val if str(x).strip()]
    s = str(val).strip()
    if not s:
        return []
    parts = []
    for chunk in s.replace(";", ",").replace("/", ",").split(","):
        chunk = chunk.strip()
        if chunk:
            parts.append(chunk)
    return parts

def extract_tech_lines_nt(tech_stack: Dict[str, Any]) -> str:
    if not isinstance(tech_stack, dict):
        tech_stack = {}
    lower_map = {str(k).lower(): v for k, v in tech_stack.items()}
    languages = normalize_list(lower_map.get("language") or lower_map.get("languages"))
    os_list   = normalize_list(lower_map.get("os") or lower_map.get("platform") or lower_map.get("operating_system"))
    tools     = normalize_list(lower_map.get("tools") or lower_map.get("tool"))
    lines = []
    if languages: lines.append(f"* language: {', '.join(languages)}")
    if os_list:   lines.append(f"* os: {', '.join(os_list)}")
    if tools:     lines.append(f"* tools: {', '.join(tools)}")
    return strip_markers("\n".join(lines))  # ← 마커 제거

def bullet_lines(items: List[str]) -> str:
    items = [str(i).strip() for i in items if str(i).strip()]
    return "\n".join(f"* {i}" for i in items)

def related_task_names_nt(related_tasks: List[Dict[str, Any]], task_id_to_name: Dict[str, str]) -> List[str]:
    names = []
    for rt in related_tasks or []:
        name = (rt.get("task_name") or "").strip()
        if not name:
            tid = (rt.get("task_id") or "").strip()
            if tid and tid in task_id_to_name:
                name = task_id_to_name[tid]
        if name:
            names.append(name)
    return names

def build_workbook_nontrack(template_bytes: bytes, org: str, role: str, data: Dict[str, Any]) -> BytesIO:
    """템플릿 서식 유지, 값만 주입"""
    wb = load_workbook(BytesIO(template_bytes))
    ws_task  = wb["Task"] if "Task" in wb.sheetnames else wb[wb.sheetnames[0]]
    ws_skill = wb["Skill"] if "Skill" in wb.sheetnames else wb[wb.sheetnames[1]]

    # Task
    set_text(ws_task, "B1", org)
    set_text(ws_task, "B2", role)
    tasks = collect_tasks_nt(data)
    task_id_to_name = {}
    for t in tasks:
        tid = str(t.get("task_id") or "").strip()
        tname = str(t.get("task_name") or "").strip()
        if tid and tname:
            task_id_to_name[tid] = tname
    row = TASK_START_ROW_NT
    for t in tasks[: (TASK_END_ROW_NT - TASK_START_ROW_NT + 1) ]:
        set_text(ws_task, f"A{row}", str(t.get("task_name") or "").strip())
        set_text(ws_task, f"C{row}", str(t.get("task_description") or "").strip())
        row += 1
    for r in range(row, TASK_END_ROW_NT + 1):
        set_text(ws_task, f"A{r}", ""); set_text(ws_task, f"C{r}", "")

    # Skill
    set_text(ws_skill, "B1", org)
    set_text(ws_skill, "B2", role)
    processed = 0
    max_rows = SKILL_END_ROW_NT - SKILL_START_ROW_NT + 1
    for s in iter_skills_nt(data):
        if processed >= max_rows: break
        r = SKILL_START_ROW_NT + processed
        rel_names = related_task_names_nt(s.get("related_tasks"), task_id_to_name)
        set_text(ws_skill, f"A{r}", bullet_lines(rel_names) if rel_names else "")
        set_text(ws_skill, f"B{r}", str(s.get("name") or "").strip())
        set_text(ws_skill, f"D{r}", strip_markers(s.get("definition")))
        set_text(ws_skill, f"F{r}", extract_tech_lines_nt(s.get("tech_stack")))
        processed += 1
    for r in range(SKILL_START_ROW_NT + processed, SKILL_END_ROW_NT + 1):
        for c in ("A","B","D","F"):
            set_text(ws_skill, f"{c}{r}", "")

    bio = BytesIO(); wb.save(bio); bio.seek(0); return bio

def process_uploaded_txt_nontrack(uploaded_file, template_bytes: bytes):
    org, role_display, role_for_filename = parse_org_role_from_filename_nt(uploaded_file.name)
    safe_org  = sanitize_filename_component(org, "org")
    safe_role = sanitize_filename_component(role_for_filename, "role")
    out_name = f"Non Track_Paper Interview_{safe_org}_{safe_role}.xlsx"
    data = load_json_from_txt_bytes(uploaded_file.read())
    wb_bytes = build_workbook_nontrack(template_bytes, org, role_display, data)
    return out_name, wb_bytes

# ==========================
# Track 파서/로직
# ==========================
def parse_org_and_job_from_filename_track(filename: str) -> Tuple[str, str]:
    """
    파일명에서 상위조직/직무:
    - {상위조직} = '_' split 첫 토큰
    - {직무} = 첫 토큰 제외 후, 끝에서 'skill'/'HC 제외' 제거, 나머지를 '_'로 결합(원문 규칙 유지)
    """
    stem = Path(filename).stem
    tokens = stem.split("_")
    if not tokens:
        return "", ""
    org = tokens[0].strip()

    def norm(t: str) -> str: return t.lower().replace(" ", "")
    tail = tokens[1:]
    while tail and norm(tail[-1]) in ("skill", "hc제외"):
        tail.pop()
    job = "_".join(tail).strip()
    return org, job

# ---- 트랙 유틸 ----
def ensure_wrap(ws, row: int, col: int, vertical: str = "center"):
    existing = ws.cell(row=row, column=col).alignment or Alignment()
    ws.cell(row=row, column=col).alignment = Alignment(
        horizontal=existing.horizontal,
        vertical=vertical,
        wrap_text=True,
        text_rotation=existing.text_rotation,
        shrink_to_fit=existing.shrink_to_fit,
        indent=existing.indent
    )

def ensure_merge(ws, cell_range: str):
    existing = {str(rng) for rng in ws.merged_cells.ranges}
    if cell_range not in existing:
        ws.merge_cells(cell_range)

def set_vertical_center_all(ws):
    max_r, max_c = ws.max_row, ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
        for cell in row:
            a = cell.alignment or Alignment()
            cell.alignment = Alignment(
                horizontal=a.horizontal,
                vertical="center",
                wrap_text=a.wrap_text,
                text_rotation=a.text_rotation,
                shrink_to_fit=a.shrink_to_fit,
                indent=a.indent
            )

def copy_sheet_by_template(wb, template_sheet_name: str, new_title: str):
    src = wb[template_sheet_name]
    new_ws = wb.copy_worksheet(src)
    new_ws.title = new_title
    # column widths
    for key, dim in src.column_dimensions.items():
        new_ws.column_dimensions[key].width = dim.width
    # row heights
    for idx, dim in src.row_dimensions.items():
        if dim.height:
            new_ws.row_dimensions[idx].height = dim.height
    # merges
    src_merges = [str(r) for r in src.merged_cells.ranges]
    new_merges = {str(r) for r in new_ws.merged_cells.ranges}
    for r in src_merges:
        if r not in new_merges:
            new_ws.merge_cells(r)
    return new_ws

# ---- 트랙 데이터 선택 ----
def select_tasks_for_track(all_tasks: List[Dict[str, Any]], track_name: str, limit: int) -> List[Dict[str, Any]]:
    sel = [t for t in (all_tasks or []) if ((t.get("track") or {}).get("name")) == track_name]
    return sel[:limit]

def get_skill_field(s: Dict[str, Any], key: str, default=None):
    """스킬 항목이 {'skill': {...}} 또는 평평한 dict 모두 지원"""
    if isinstance(s, dict) and "skill" in s and isinstance(s["skill"], dict):
        return s["skill"].get(key, default)
    return s.get(key, default)

def get_skill_related_tasks(s: Dict[str, Any]):
    if isinstance(s, dict) and "skill" in s:
        return s.get("related_tasks") or s["skill"].get("related_tasks") or []
    return s.get("related_tasks") or []

def get_skill_track(s: Dict[str, Any]) -> Dict[str, Any]:
    # 주로 최상위에 'track'이 온다고 가정
    return s.get("track") or {}

def select_skills_for_track(all_skills: List[Dict[str, Any]], track_name: str, track_code: str, limit: int) -> List[Dict[str, Any]]:
    matched = []
    for s in all_skills or []:
        tr = get_skill_track(s) or {}
        scope = s.get("track_scope")
        name_match = (tr.get("name") == track_name) or (tr.get("code") == track_code)
        if name_match:
            matched.append(s); continue
        if scope == "common":
            for rt in get_skill_related_tasks(s) or []:
                trt = (rt.get("track") or {})
                if (trt.get("name") == track_name) or (trt.get("code") == track_code):
                    matched.append(s); break
    # 중복 제거(스킬명 기준)
    uniq, seen = [], set()
    for s in matched:
        sk_name = (get_skill_field(s, "name") or "").strip()
        if sk_name and sk_name not in seen:
            seen.add(sk_name); uniq.append(s)
    # rank 오름차순, None은 뒤
    def rank_key(s):
        r = get_skill_field(s, "rank")
        return (r is None, r if r is not None else 10**9)
    uniq.sort(key=rank_key)
    return uniq[:limit]

# ---- 트랙 본문 가공 ----
def bullets_from_related_tasks(related_tasks: List[Dict[str, Any]], current_track_name: str) -> str:
    if not related_tasks: return ""
    names, seen = [], set()
    for rt in related_tasks:
        tname = (rt or {}).get("task_name")
        ttrack = ((rt or {}).get("track") or {}).get("name")
        if tname and (ttrack == current_track_name) and (tname not in seen):
            seen.add(tname); names.append(tname)
    return "\n".join(f"* {n}" for n in names)

def listify_tech_value(v) -> List[str]:
    if v is None: return []
    if isinstance(v, (list, tuple, set)):
        return [strip_markers(x) for x in v if str(x).strip()]
    # 문자열이면 구분자로 분리
    return [strip_markers(x.strip()) for x in re.split(r"[;,/]", str(v)) if x.strip()]

def bullets_from_tech_stack(tech_stack: Dict[str, Any]) -> str:
    tech_stack = tech_stack or {}
    lines = []
    for key in ("language", "os", "tools"):
        vals = tech_stack.get(key)
        items = listify_tech_value(vals)
        items = [x for x in items if x]  # 빈 문자열 제거
        if items:
            lines.append(f"* {key}: {', '.join(items)}")
    return "\n".join(lines)

# ---- 트랙 시트 쓰기 ----
def write_task_sheet(ws, org_name: str, job_name: str, track_name: str, tasks: List[Dict[str, Any]]):
    ws["B1"].value = org_name
    ws["B2"].value = job_name
    ensure_merge(ws, TRACK_TITLE_RANGE_T)
    ws["D1"].value = track_name
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = TASK_ROW_START_T
    for t in tasks:
        if row > TASK_ROW_END_T: break
        ws.cell(row=row, column=1).value = t.get("task_name") or ""
        desc = t.get("task_description") or ""
        ws.cell(row=row, column=3).value = desc
        ensure_wrap(ws, row, 3, vertical="center")
        row += 1
    set_vertical_center_all(ws)

def write_skill_sheet(ws, org_name: str, job_name: str, track_name: str, skills: List[Dict[str, Any]]):
    ws["B1"].value = org_name
    ws["B2"].value = job_name
    ensure_merge(ws, TRACK_TITLE_RANGE_T)
    ws["D1"].value = track_name
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = SKILL_ROW_START_T
    for s in skills:
        if row > SKILL_ROW_END_T: break
        # A: 유관업무(현재 트랙 기준)
        a_text = bullets_from_related_tasks(get_skill_related_tasks(s), current_track_name=track_name)
        ws.cell(row=row, column=1).value = a_text
        ensure_wrap(ws, row, 1, vertical="center")
        # B: 스킬명
        ws.cell(row=row, column=2).value = (get_skill_field(s, "name") or "")
        # D: 설명(마커 제거)
        d_text = strip_markers(get_skill_field(s, "definition"))
        ws.cell(row=row, column=4).value = d_text
        ensure_wrap(ws, row, 4, vertical="center")
        # F: tech_stack(language/os/tools) (마커 제거 포함)
        f_text = bullets_from_tech_stack(get_skill_field(s, "tech_stack") or {})
        ws.cell(row=row, column=6).value = f_text
        ensure_wrap(ws, row, 6, vertical="center")
        row += 1
    set_vertical_center_all(ws)

def build_workbook_track(template_bytes: bytes, org: str, job: str, data: Dict[str, Any]) -> BytesIO:
    wb = load_workbook(BytesIO(template_bytes))

    # 트랙 목록(meta.tracks 우선)
    tracks = []
    meta_tracks = (((data.get("meta") or {}).get("tracks")) or [])
    if meta_tracks:
        for idx, tr in enumerate(meta_tracks, start=1):
            tracks.append({"index": idx, "name": tr.get("track_name"), "code": tr.get("track_code")})
    else:
        seen, idx = set(), 1
        for t in data.get("tasks", []):
            tn = (t.get("track") or {}).get("name")
            tc = (t.get("track") or {}).get("code")
            if tn and (tn, tc) not in seen:
                tracks.append({"index": idx, "name": tn, "code": tc})
                seen.add((tn, tc)); idx += 1

    all_tasks  = data.get("tasks")  or []
    all_skills = data.get("skills") or []

    for tr in tracks:
        t_idx = tr["index"]; t_name = tr["name"]; t_code = tr.get("code")
        # Task 시트
        task_ws_title = f"트랙 {t_idx}_Task"
        task_ws = copy_sheet_by_template(wb, TASK_TEMPLATE_SHEET_T, task_ws_title)
        tasks_for_track = select_tasks_for_track(all_tasks, t_name, limit=(TASK_ROW_END_T - TASK_ROW_START_T + 1))
        write_task_sheet(task_ws, org_name=org, job_name=job, track_name=t_name, tasks=tasks_for_track)
        # Skill 시트
        skill_ws_title = f"트랙 {t_idx}_Skill"
        skill_ws = copy_sheet_by_template(wb, SKILL_TEMPLATE_SHEET_T, skill_ws_title)
        skills_for_track = select_skills_for_track(all_skills, t_name, t_code, limit=(SKILL_ROW_END_T - SKILL_ROW_START_T + 1))
        write_skill_sheet(skill_ws, org_name=org, job_name=job, track_name=t_name, skills=skills_for_track)

    # 원본 템플릿 Task/Skill 시트 제거(Description 등은 유지)
    for base in (TASK_TEMPLATE_SHEET_T, SKILL_TEMPLATE_SHEET_T):
        if base in wb.sheetnames:
            wb.remove(wb[base])

    bio = BytesIO(); wb.save(bio); bio.seek(0); return bio

def process_uploaded_txt_track(uploaded_file, template_bytes: bytes):
    org, job = parse_org_and_job_from_filename_track(uploaded_file.name)
    safe_org = sanitize_filename_component(org, "org")
    safe_job = sanitize_filename_component(job, "job")
    out_name = f"Track_Paper Interview_{safe_org}_{safe_job}.xlsx"
    data = load_json_from_txt_bytes(uploaded_file.read())
    wb_bytes = build_workbook_track(template_bytes, org, job, data)
    return out_name, wb_bytes

# ==========================
# 순차(멀티) 다운로드
# ==========================
def render_sequential_downloads(results_bytes: Dict[str, bytes], height: int = 240):
    if not results_bytes:
        return
    items_html, hidden_links_html = [], []
    for fname, b in results_bytes.items():
        b64 = base64.b64encode(b).decode("utf-8")
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        data_uri = f"data:{mime};base64,{b64}"
        items_html.append(f"<li>{fname}</li>")
        hidden_links_html.append(
            f'<a class="dl-link" href="{data_uri}" download="{fname}" style="display:none;"></a>'
        )
    html = f"""
<div id="bulk-dl">
  <button id="btn-bulk" style="padding:0.6rem 1rem;font-size:1rem;">📥 전체 파일 순차 다운로드</button>
  <p style="margin:0.5rem 0 0.25rem 0;">브라우저에서 다중 다운로드 허용이 필요할 수 있습니다.</p>
  <ul style="margin-top:0.25rem;">{''.join(items_html)}</ul>
  {''.join(hidden_links_html)}
</div>
<script>
(function() {{
  const btn = document.getElementById('btn-bulk');
  btn.addEventListener('click', async () => {{
    const links = Array.from(document.querySelectorAll('#bulk-dl a.dl-link'));
    for (const a of links) {{
      a.click();
      await new Promise(r => setTimeout(r, 400));
    }}
  }});
}})();
</script>
"""
    components.html(html, height=height, scrolling=False)

# ==========================
# Streamlit UI
# ==========================
st.set_page_config(page_title="TXT → Excel 변환기 (Non Track / Track)", layout="wide")
st.title("TXT(JSON) → Excel 변환기")

mode = st.radio("모드 선택", options=["Non Track", "Track"], horizontal=True)

# 템플릿 선택
with st.sidebar:
    st.header("템플릿 설정")
    tpl_upload = st.file_uploader("템플릿 업로드 (.xlsx) — (선택)", type=["xlsx"], accept_multiple_files=False)

    if mode == "Non Track":
        default_tpl_path = TEMPLATE_DIR / DEFAULT_TEMPLATE_NONTRACK
        tpl_label = DEFAULT_TEMPLATE_NONTRACK
    else:
        default_tpl_path = TEMPLATE_DIR / DEFAULT_TEMPLATE_TRACK
        tpl_label = DEFAULT_TEMPLATE_TRACK

    if tpl_upload is None:
        if not default_tpl_path.exists():
            st.error(f"기본 템플릿이 없습니다: {default_tpl_path}")
        else:
            st.success(f"기본 템플릿 사용: {tpl_label}")
            template_bytes = default_tpl_path.read_bytes()
    else:
        template_bytes = tpl_upload.read()
        st.success(f"업로드한 템플릿 사용: {tpl_upload.name}")

    st.divider()
    if mode == "Non Track":
        st.markdown("**고정 옵션 (Non Track)**")
        st.markdown("- 기존 **열너비/행높이/서식** 유지")
        st.markdown("- **줄바꿈** 표시(wrap_text=True)")
        st.markdown("- 스킬 설명/테크 스택의 **[cite: …], (Source …)** 제거")
    else:
        st.markdown("**고정 옵션 (Track)**")
        st.markdown("- 템플릿 시트(Task/Skill) 복제 → 트랙별 시트 생성")
        st.markdown("- `D1:D2` 병합 및 트랙명 표시")
        st.markdown("- 기존 **열너비/행높이/서식** 유지")
        st.markdown("- 스킬 설명/테크 스택의 **[cite: …], (Source …)** 제거")

# 업로더
st.subheader("1) TXT(JSON) 파일 업로드")
uploaded_files = st.file_uploader("여러 파일을 동시에 올릴 수 있습니다.", type=["txt"], accept_multiple_files=True)

# 미리보기
if uploaded_files:
    st.write("**파일명 파싱 미리보기**")
    preview = []
    for f in uploaded_files:
        if mode == "Non Track":
            org, role_display, role_for_filename = parse_org_role_from_filename_nt(f.name)
            out = f"Non Track_Paper Interview_{sanitize_filename_component(org)}_{sanitize_filename_component(role_for_filename)}.xlsx"
            preview.append({"원본 파일": f.name, "상위조직명": org, "직무명": role_display, "생성될 엑셀": out})
        else:
            org, job = parse_org_and_job_from_filename_track(f.name)
            out = f"Track_Paper Interview_{sanitize_filename_component(org)}_{sanitize_filename_component(job)}.xlsx"
            preview.append({"원본 파일": f.name, "상위조직명": org, "직무명(파일 규칙)": job, "생성될 엑셀": out})
    st.dataframe(preview, use_container_width=True)

run = st.button("변환 실행", type="primary", disabled=not uploaded_files)

# 세션 상태 (다운로드 후에도 결과 유지)
if "results_data" not in st.session_state:
    st.session_state["results_data"] = {}
if "errors_data" not in st.session_state:
    st.session_state["errors_data"] = []
if "last_mode" not in st.session_state:
    st.session_state["last_mode"] = mode

if run and uploaded_files:
    if "template_bytes" not in locals():
        st.error("템플릿을 찾을 수 없습니다. 사이드바에서 템플릿을 업로드하거나 기본 템플릿을 확인하세요.")
    else:
        results: Dict[str, bytes] = {}
        errors: List[str] = []
        with st.spinner("변환 중..."):
            for uf in uploaded_files:
                try:
                    if mode == "Non Track":
                        name, bio = process_uploaded_txt_nontrack(uf, template_bytes)
                    else:
                        name, bio = process_uploaded_txt_track(uf, template_bytes)
                    results[name] = bio.getvalue()
                except Exception as e:
                    errors.append(f"{uf.name} → 실패: {e}")
        st.session_state["results_data"] = results
        st.session_state["errors_data"] = errors
        st.session_state["last_mode"] = mode

# 결과 렌더(세션 유지)
results_data: Dict[str, bytes] = st.session_state.get("results_data", {})
errors_data: List[str] = st.session_state.get("errors_data", [])
last_mode = st.session_state.get("last_mode", mode)

if results_data:
    st.subheader("2) 변환 결과")
    col1, col2 = st.columns([2, 1])

    with col1:
        st.success(f"{len(results_data)}개 파일 생성 완료 — 모드: {last_mode}")
        for fname, b in results_data.items():
            st.download_button(
                label=f"⬇️ {fname} 다운로드",
                data=b,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    with col2:
        render_sequential_downloads(results_data)

if errors_data:
    st.warning("일부 파일 변환 중 오류가 발생했습니다.")
    for msg in errors_data:
        st.write(f"• {msg}")

st.divider()
if mode == "Non Track":
    st.markdown(
        """
**규칙 요약 — Non Track**
- 파일명  
  - `{상위조직명}` = `_` 분할 첫 토큰  
  - `{직무명}` = 두 번째 토큰부터, 끝에서 `'skill'`, `'HC 제외'` 제거 → 공백 연결  
- Task 시트  
  - `B1={상위조직명}`, `B2={직무명}`  
  - `A5..A14 = tasks[*].task_name`, `C5..C14 = tasks[*].task_description`  
- Skill 시트  
  - `B1={상위조직명}`, `B2={직무명}`  
  - `A5..A11 = related_tasks[*].task_name`을 `* 항목` 줄바꿈 목록  
  - `B5..B11 = skill.name`, `D5..D11 = skill.definition(마커 제거)`, `F5..F11 = tech_stack(language/os/tools, 마커 제거)`  
- 템플릿 **열너비/행높이/서식 유지**, 줄바꿈 표시(wrap_text=True)
        """
    )
else:
    st.markdown(
        """
**규칙 요약 — Track**
- 파일명  
  - `{상위조직명}` = `_` 분할 첫 토큰  
  - `{직무명}` = 첫 토큰 제외 후, 끝에서 `'skill'`, `'HC 제외'` 제거 → **`_`로 결합**  
- 트랙 처리  
  - `meta.tracks`가 있으면 이를 사용, 없으면 `tasks[*].track`에서 유추  
  - 트랙별 시트 생성: **`트랙 n_Task`**, **`트랙 n_Skill`** (템플릿 Task/Skill 복제)  
  - `D1:D2` 병합 + 트랙명 표시  
- Task(트랙별)  
  - `B1={상위조직명}`, `B2={직무명}`  
  - `A5..A14 = 해당 트랙의 tasks[*].task_name`  
  - `C5..C14 = 해당 트랙의 tasks[*].task_description`  
- Skill(트랙별)  
  - 스킬 선택:  
    1) `skill.track.name/code == 현재 트랙` 포함  
    2) `track_scope == "common"`이고 related_tasks에 현재 트랙 연결이 있으면 포함  
    - 스킬명 중복 제거, `rank` 오름차순(없으면 뒤)로 정렬, 행 수 제한  
  - `A5..A11 = 현재 트랙의 related_tasks[*].task_name`만 bullet  
  - `B5..B11 = skill.name`  
  - `D5..D11 = skill.definition(마커 제거)`  
  - `F5..F11 = tech_stack(language/os/tools, 마커 제거)`  
- 템플릿 **열너비/행높이/서식 유지**, 줄바꿈 표시(wrap_text=True)
        """
    )
